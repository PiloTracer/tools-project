from __future__ import annotations

import uuid
from datetime import date
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db import get_db
from app.deps import get_current_user
from app.models.client import Client
from app.models.client_contact import ClientContact
from app.models.project import Project
from app.models.prospect import PIPELINE_STAGE_ORDER, Prospect
from app.models.task import Task
from app.models.ticket import Ticket
from app.models.user import User
from app.services.project_access import require_project_access

router = APIRouter(prefix="/v1/reports", tags=["reports"])

STAGE_LABELS: dict[str, str] = {
    "target": "Target",
    "connected": "Connected",
    "engaged": "Engaged",
    "call_scheduled": "Call Scheduled",
    "call_done": "Call Done",
    "proposal_sent": "Proposal Sent",
    "negotiating": "Negotiating",
    "won": "Won",
    "lost": "Lost",
}


def _build_workbook(headers: list[str], rows: list[list]) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2B3A5C", end_color="2B3A5C", fill_type="solid")
    cell_font = Font(name="Calibri", size=10)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, len(rows) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

    ws.freeze_panes = "A2"
    wb.save("/tmp/_report.xlsx")
    with open("/tmp/_report.xlsx", "rb") as f:
        return f.read()


@router.get("/pipeline")
async def pipeline_report(
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    stage: str | None = None,
    source: str | None = None,
):
    q = select(Prospect).order_by(Prospect.created_at.desc())
    if stage:
        q = q.where(Prospect.pipeline_stage == stage)
    if source:
        q = q.where(Prospect.source == source)
    rows = (await db.scalars(q)).all()

    headers = [
        "Company Name", "Pipeline Stage", "Value (USD)", "Source",
        "First Contact", "Last Interaction", "Next Action", "Next Action Date",
        "Notes", "Created At",
    ]
    data = []
    for p in rows:
        data.append([
            p.company_name,
            STAGE_LABELS.get(p.pipeline_stage, p.pipeline_stage),
            float(p.pipeline_value) if p.pipeline_value else "",
            p.source or "",
            str(p.first_contact_date) if p.first_contact_date else "",
            p.last_interaction.strftime("%Y-%m-%d %H:%M") if p.last_interaction else "",
            p.next_action or "",
            str(p.next_action_date) if p.next_action_date else "",
            p.notes or "",
            p.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    excel_bytes = _build_workbook(headers, data)
    return StreamingResponse(
        iter([excel_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pipeline-report.xlsx"},
    )


@router.get("/clients")
async def clients_report(
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    q = select(Client).order_by(Client.name)
    clients = (await db.scalars(q)).all()

    contact_rows = (await db.scalars(select(ClientContact))).all()
    contacts_by_client: dict[uuid.UUID, list[ClientContact]] = {}
    for ct in contact_rows:
        contacts_by_client.setdefault(ct.client_id, []).append(ct)

    headers = [
        "Client Name", "Slug", "Industry", "Notes",
        "Contact Name", "Contact Email", "Contact Phone", "Contact Title", "Contact Role",
        "Created At",
    ]
    data = []
    for c in clients:
        contacts = contacts_by_client.get(c.id, [])
        if contacts:
            for ct in contacts:
                data.append([
                    c.name, c.slug, c.industry or "", c.notes or "",
                    ct.name, ct.email, ct.phone or "", ct.title or "", ct.role,
                    c.created_at.strftime("%Y-%m-%d"),
                ])
        else:
            data.append([
                c.name, c.slug, c.industry or "", c.notes or "",
                "", "", "", "", "",
                c.created_at.strftime("%Y-%m-%d"),
            ])

    excel_bytes = _build_workbook(headers, data)
    return StreamingResponse(
        iter([excel_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=clients-report.xlsx"},
    )


@router.get("/projects/{project_id}/tasks")
async def project_tasks_report(
    project_id: uuid.UUID,
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    status: str | None = None,
):
    await require_project_access(db, user, project_id)
    project = await db.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    q = select(Task).where(Task.project_id == project_id).order_by(Task.created_at.desc())
    if status:
        q = q.where(Task.status == status)
    tasks = (await db.scalars(q)).all()

    headers = [
        "Ref", "Title", "Status", "Priority", "Assignee ID",
        "Due At", "Closed At", "Created At",
    ]
    data = []
    for t in tasks:
        data.append([
            t.ref or "",
            t.title,
            t.status,
            t.priority,
            str(t.assignee_id) if t.assignee_id else "",
            t.due_at.strftime("%Y-%m-%d") if t.due_at else "",
            t.closed_at.strftime("%Y-%m-%d %H:%M") if t.closed_at else "",
            t.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    excel_bytes = _build_workbook(headers, data)
    return StreamingResponse(
        iter([excel_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={project.slug}-tasks.xlsx"},
    )


@router.get("/projects/{project_id}/tickets")
async def project_tickets_report(
    project_id: uuid.UUID,
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    status: str | None = None,
    queue_slug: str | None = None,
):
    await require_project_access(db, user, project_id)
    project = await db.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    q = select(Ticket).where(Ticket.project_id == project_id).order_by(Ticket.created_at.desc())
    if status:
        q = q.where(Ticket.status == status)
    if queue_slug:
        q = q.where(Ticket.queue_slug == queue_slug)
    tickets = (await db.scalars(q)).all()

    headers = [
        "Ref", "Title", "Status", "Priority", "Queue", "Requester Email",
        "Assignee ID", "First Response", "Resolved At", "Closed At", "Created At",
    ]
    data = []
    for t in tickets:
        data.append([
            t.ref or "",
            t.title,
            t.status,
            t.priority,
            t.queue_slug,
            t.requester_email or "",
            str(t.assignee_id) if t.assignee_id else "",
            t.first_response_at.strftime("%Y-%m-%d %H:%M") if t.first_response_at else "",
            t.resolved_at.strftime("%Y-%m-%d %H:%M") if t.resolved_at else "",
            t.closed_at.strftime("%Y-%m-%d %H:%M") if t.closed_at else "",
            t.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    excel_bytes = _build_workbook(headers, data)
    return StreamingResponse(
        iter([excel_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={project.slug}-tickets.xlsx"},
    )
